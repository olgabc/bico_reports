import os
import openpyxl
import pandas as pd
from libs.db_libs.write import write_some_xlsx


def load_some_xlsx(name, folder, index_col=None):
    file_path = os.path.join(
        folder,
        name
    )

    some_dataframe = pd.read_excel(
        file_path,
        encoding="cp1251",
        sep=";",
        index_col=index_col
    )
    return some_dataframe


def load_some_xlsx_with_hyperlinks(name, folder, col_nums):
    df = load_some_xlsx(name, folder)
    df = df.drop(
        ["#",
         "Номер закупки",
         "Названия позиций",
         "Участник",
         "ИНН участника",
         "Статус участника",
         "Контракт"
         ],
        axis=1)
    file_path = os.path.join(
        folder,
        name
    )
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name('List 1')
    max_row = len(df)

    for col_num in col_nums:
        links = []

        for i in range(2, max_row+2):  # 2nd arg in range() not inclusive, so add 1

            try:
                links.append(ws.cell(row=i, column=col_num).hyperlink.target)
                print(ws.cell(row=i, column=col_num).hyperlink.target)

            except AttributeError:
                links.append("")
                print("p", ws.cell(row=i, column=col_num).value)

        links_se = pd.Series(links)
        col_name = "link_{}".format(col_num)
        df[col_name] = links_se.values

    return df


def load_some_csv(name, folder):
    file_path = os.path.join(
        folder,
        name
    )

    some_dataframe = pd.read_csv(
        file_path,
        encoding="cp1251",
        sep=";",
        engine="python"
    )
    return some_dataframe

#print(load_some_xlsx("_пример.xlsx", r'c:\bico\test_xlsx'))