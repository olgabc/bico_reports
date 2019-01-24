import os
import openpyxl
from config.config import project_folder


def get_hyperlinks_adresses_col(
    xls_file,
    xls_list,
    link_col,
    folder="",

):
    """

    :param xls_file: xls_file containing column with hyperlinks
    :param xls_list: xls_list containing column with hyperlinks
    :param link_col: number of column with hyperlinks
    :param folder: folder containing xls_file
    :return: list of hyperlinks_adresses
    """

    file_path = os.path.join(
        project_folder,
        folder,
        xls_file
    )

    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(xls_list)
    max_row = ws.max_row

    links_list = []

    for i in range(2, max_row + 1):

        try:
            link = ws.cell(row=i, column=link_col).hyperlink.target
            links_list.append(link)

            print("link adress:", link)

        except AttributeError:
            links_list.append("")
            print("hyperlink text:", ws.cell(row=i, column=link_col).value)

    return links_list
