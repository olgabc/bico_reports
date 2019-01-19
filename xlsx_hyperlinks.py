import os
from string import ascii_uppercase as alphabet
import openpyxl
import xlsxwriter

excel_cols = []

for letter in alphabet:
    excel_cols.append(letter)

for first_letter in alphabet:
    for second_letter in alphabet:
        excel_cols.append(first_letter + second_letter)

for first_letter in alphabet:
    for second_letter in alphabet:
        for third_letter in alphabet:
            excel_cols.append(first_letter + second_letter + third_letter)

excel_cols_dict = {i+1: excel_cols[i] for i in range(len(excel_cols))}


def get_hyperlinks_adresses_col(
    xls_file,
    xls_list,
    link_col,
    folder="",

):
    """

    :param xls_file: xls_file containing column with hyperlinks
    :param xls_list: xls_list containing column with hyperlinks
    :param link_col: number of column with hyperlinks
    :param folder: folder containing xls_file
    :return: list of hyperlinks_adresses
    """

    file_path = os.path.join(
        folder,
        xls_file
    )

    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(xls_list)
    max_row = ws.max_row

    links_list = []

    for i in range(2, max_row + 1):

        try:
            link = ws.cell(row=i, column=link_col).hyperlink.target
            links_list.append(link)

            print("link adress:", link)

        except AttributeError:
            links_list.append("")
            print("hyperlink text:", ws.cell(row=i, column=link_col).value)

    return links_list


def set_hyperlinlks_in_excel_col(
    hyperlinks_adresses,
    hyperlinks_texts=None,
    xl_col=None,
    tip="",
    preffix="",
    xls_file="hyperlinks.xlsx",
    xls_worksheet="hyperlinks",
    folder="",
):
    """
    :param hyperlinks_adresses:
    :param hyperlinks_texts:
    :param xl_col:
    :param tip:
    :param preffix:
    :param xls_file:
    :param xls_worksheet:
    :param folder:
    :return: None
    """
    max_row = len(hyperlinks_adresses) + 1

    if preffix:
        hyperlinks_adresses = ["{}/{}/".format(preffix, hyperlink_adress) for hyperlink_adress in hyperlinks_adresses]

    xls_file_path = os.path.abspath(os.path.join(folder, xls_file))
    workbook = xlsxwriter.Workbook(xls_file_path)
    worksheet = workbook.add_worksheet(xls_worksheet)

    for xl_row, link_adress, link_text in zip(range(1, max_row), hyperlinks_adresses, hyperlinks_texts):
        print("link:", link_adress, "text:", link_text)
        worksheet.write_url(
            row=xl_row,
            col=xl_col,
            url=link_adress,
            string=link_text,
            tip=tip
        )

    workbook.close()
